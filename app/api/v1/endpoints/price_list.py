import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Security
from fastapi.security import APIKeyHeader
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any

from app.dependencies import get_db
from app.models.item import Item as ItemModel
from app.core.config import settings

router = APIRouter(
    prefix="/price-list",
    tags=["Price List"],
)

# --- Безопасность (Подтверждено: Оставляем токен) ---
API_KEY_NAME = "X-Admin-Token"
api_key_header = APIKeyHeader(name=API_KEY_NAME, auto_error=False)

async def get_admin_user(api_key: str = Security(api_key_header), db: Session = Depends(get_db)):
    """Проверяет, совпадает ли токен из заголовка с токеном в .env"""
    if api_key == settings.ADMIN_API_TOKEN and settings.ADMIN_API_TOKEN != "your_super_secret_api_token_12345":
        return True
    
    # 💡 Добавил проверку, что токен не является дефолтным
    if api_key != settings.ADMIN_API_TOKEN:
         raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Неверный или отсутствующий Admin API Token"
        )
    # Если используется дефолтный токен
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="API Token не настроен. Пожалуйста, измените ADMIN_API_TOKEN в .env"
    )

# --------------------


@router.get("/download", dependencies=[Depends(get_admin_user)])
async def download_price_list(db: Session = Depends(get_db)):
    """
    Генерирует и отдает Excel-файл со всеми вариантами товаров.
    """
    
    # ... (логика получения данных и создания файла)
    items = db.query(ItemModel).order_by(ItemModel.name, ItemModel.id).all()
    # ... (создание wb, ws, заголовки, стили - БЕЗ ИЗМЕНЕНИЙ)

    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    headers = ['ID (Не менять!)', 'Название', 'Память', 'Цвет', 'Цена (Редактировать)']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    column_widths = {'A': 15, 'B': 40, 'C': 15, 'D': 15, 'E': 20}

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Заполняем данными
    for item in items:
        # Убедимся, что цена является числом для Excel
        try:
             item_price = float(item.price)
        except (TypeError, ValueError):
             item_price = 0.0
             
        row = [
            item.id,
            item.name,
            item.memory or '—',
            item.color or '—',
            item_price
        ]
        ws.append(row)


    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=price_list_exported_{db.query(ItemModel).count()}_items.xlsx"
        }
    )


@router.post("/upload", dependencies=[Depends(get_admin_user)])
async def upload_price_list(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Принимает Excel-файл, парсит его и МАССОВО обновляет цены в БД.
    """
    
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Неверный формат. Нужен .xlsx файл.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
        ws = wb.active

        updates = []
        errors = []
        
        # 2. Парсим строки
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue 

            try:
                item_id = int(row[0])
                new_price_raw = row[4] 

                # 💡 Улучшенная обработка цены из колонки 'E' (индекс 4)
                if isinstance(new_price_raw, (int, float)):
                    new_price = float(new_price_raw)
                elif isinstance(new_price_raw, str):
                    new_price = float(new_price_raw.replace(',', '.'))
                else:
                    raise ValueError("Цена не является числом или строкой.")
                
                if item_id <= 0:
                    raise ValueError("ID должен быть положительным числом.")
                # 💡 ИСПРАВЛЕНО: Цена не может быть отрицательной, но может быть 0
                if new_price < 0:
                    raise ValueError("Цена не может быть отрицательной.") 
                    
                updates.append({'id': item_id, 'price': new_price})
                
            except Exception as e:
                item_id_for_log = row[0] if row and row[0] is not None else '?' 
                errors.append(f"Строка с ID {item_id_for_log}: неверный формат. Ошибка: {str(e)[:50]}...")

        if not updates:
            raise HTTPException(status_code=400, detail="Файл не содержит валидных данных для обновления цен.")

        # 3. МАССОВОЕ ОБНОВЛЕНИЕ
        db.bulk_update_mappings(ItemModel, updates)
        db.commit()

        return {
            "status": "success",
            "updated": len(updates),
            "skipped": len(errors),
            "errors": errors if errors else None
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Критическая ошибка обработки файла: {str(e)}")